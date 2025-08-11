import tkinter as tk
from tkinter import ttk, messagebox
from models import Xe, VeThang
from auth import XacThuc
from openpyxl import Workbook
from datetime import datetime


def in_ve_xe(bien_so, loai_xe):
    # Tạo file Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Vé xe"

    # Thêm thông tin vào file Excel
    ws['A1'] = "Bãi xe Trọng Đạo"
    ws['A2'] = "Biển số xe:"
    ws['B2'] = bien_so
    ws['A3'] = "Loại xe:"
    ws['B3'] = loai_xe
    ws['A4'] = "Thời gian vào:"
    ws['B4'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Lưu file Excel
    ten_file = f"ve_xe_{bien_so}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(ten_file)
    messagebox.showinfo("Thông báo", f"Đã in vé xe: {ten_file}")


class UngDungQuanLyXe:
    def __init__(self, root):
        self.Xe = None
        self.root = root
        self.root.title("Quản lý bãi gửi xe")
        self.xac_thuc = XacThuc()
        self.hien_thi_man_hinh_dang_nhap()

    def hien_thi_man_hinh_dang_nhap(self):
        self.xoa_man_hinh()
        tk.Label(self.root, text="Tên đăng nhập:").grid(row=0, column=0)
        self.ten_dang_nhap_entry = tk.Entry(self.root)
        self.ten_dang_nhap_entry.grid(row=0, column=1)

        tk.Label(self.root, text="Mật khẩu:").grid(row=1, column=0)
        self.mat_khau_entry = tk.Entry(self.root, show="*")
        self.mat_khau_entry.grid(row=1, column=1)

        tk.Button(self.root, text="Đăng nhập", command=self.dang_nhap).grid(row=2, column=0, columnspan=2)

    def dang_nhap(self):
        ten_dang_nhap = self.ten_dang_nhap_entry.get()
        mat_khau = self.mat_khau_entry.get()
        if self.xac_thuc.dang_nhap(ten_dang_nhap, mat_khau):
            self.hien_thi_man_hinh_chinh()
        else:
            messagebox.showerror("Lỗi", "Tên đăng nhập hoặc mật khẩu không đúng.")

    def hien_thi_man_hinh_chinh(self):
        self.xoa_man_hinh()

        # Tạo khung chứa các nút chức năng bên trái
        left_frame = tk.Frame(self.root, width=200, bg="lightgray")
        left_frame.pack(side="left", fill="y")

        # Tạo khung chứa ảnh bên phải
        right_frame = tk.Frame(self.root, bg="white")
        right_frame.pack(side="right", fill="both", expand=True)
        # Thêm các nút chức năng vào khung bên trái
        tk.Button(left_frame, text="Thêm xe", command=self.hien_thi_man_hinh_them_xe, width=20).pack(pady=5)
        tk.Button(left_frame, text="Tính phí", command=self.hien_thi_man_hinh_tinh_phi, width=20).pack(pady=5)
        tk.Button(left_frame, text="Danh sách xe", command=self.hien_thi_man_hinh_danh_sach_xe, width=20).pack(pady=5)
        tk.Button(left_frame, text="Xóa xe", command=self.hien_thi_man_hinh_xoa_xe, width=20).pack(pady=5)
        tk.Button(left_frame, text="Tìm kiếm xe", command=self.hien_thi_man_hinh_tim_kiem_xe, width=20).pack(pady=5)
        tk.Button(left_frame, text="Vé tháng", command=self.hien_thi_man_hinh_quan_ly_ve_thang, width=20).pack(pady=5)
        tk.Button(left_frame, text="Lịch sử", command=self.hien_thi_man_hinh_lich_su, width=20).pack(pady=5)

    def xoa_man_hinh(self):
        for widget in self.root.winfo_children():
            widget.destroy()

    # --- THÊM XE ---
    def hien_thi_man_hinh_them_xe(self):
        self.xoa_man_hinh()
        tk.Label(self.root, text="Biển số:").grid(row=0, column=0)
        self.bien_so_entry = tk.Entry(self.root)
        self.bien_so_entry.grid(row=0, column=1)

        tk.Label(self.root, text="Loại xe:").grid(row=1, column=0)
        self.loai_xe_combobox = ttk.Combobox(self.root, values=["Xe máy", "Ô tô"])
        self.loai_xe_combobox.grid(row=1, column=1)
        self.loai_xe_combobox.current(0)

        tk.Label(self.root, text="Hình thức:").grid(row=2, column=0)
        self.loai_thanh_toan_combobox = ttk.Combobox(self.root, values=["Theo buổi", "Vé tháng"])
        self.loai_thanh_toan_combobox.grid(row=2, column=1)
        self.loai_thanh_toan_combobox.current(0)

        tk.Button(self.root, text="Thêm", command=self.them_xe).grid(row=3, column=0, columnspan=2)
        tk.Button(self.root, text="Quay lại", command=self.hien_thi_man_hinh_chinh).grid(row=4, column=0, columnspan=2)

    def them_xe(self):
        bien_so = self.bien_so_entry.get().strip()
        loai_xe = self.loai_xe_combobox.get()
        loai_thanh_toan = "buổi" if self.loai_thanh_toan_combobox.get() == "Theo buổi" else "thang"

        if not bien_so:
            messagebox.showerror("Lỗi", "Vui lòng nhập biển số!")
            return

        ket_qua = Xe.them_xe(bien_so, loai_xe, loai_thanh_toan)
        if ket_qua == "Thêm xe thành công.":
            messagebox.showinfo("Thành công", ket_qua)
            # In vé xe
            in_ve_xe(bien_so, loai_xe)
            self.hien_thi_man_hinh_chinh()
        else:
            messagebox.showerror("Lỗi", ket_qua)

    # --- TÍNH PHÍ ---
    def hien_thi_man_hinh_tinh_phi(self):
        self.xoa_man_hinh()
        tk.Label(self.root, text="Nhập ID/biển số:").grid(row=0, column=0)
        self.id_bien_so_entry = tk.Entry(self.root)
        self.id_bien_so_entry.grid(row=0, column=1)

        tk.Button(self.root, text="Tính phí", command=self.tinh_phi).grid(row=1, column=0, columnspan=2)
        tk.Button(self.root, text="Quay lại", command=self.hien_thi_man_hinh_chinh).grid(row=2, column=0, columnspan=2)

    def tinh_phi(self):
        id_bien_so = self.id_bien_so_entry.get().strip()
        if not id_bien_so:
            messagebox.showerror("Lỗi", "Vui lòng nhập ID hoặc biển số!")
            return

        ket_qua = Xe.tinh_phi(id_bien_so)
        messagebox.showinfo("Kết quả", ket_qua)
        self.hien_thi_man_hinh_chinh()

    # --- DANH SÁCH XE ---
    def hien_thi_man_hinh_danh_sach_xe(self):
        self.xoa_man_hinh()
        danh_sach = Xe.danh_sach_xe()

        if not danh_sach:
            tk.Label(self.root, text="Không có xe trong bãi").pack()
        else:
            for idx, xe in enumerate(danh_sach):
                tk.Label(self.root, text=f"ID: {xe[0]} | Biển: {xe[1]} | Loại: {xe[2]} | Vào: {xe[3]}").grid(row=idx, column=0)

        tk.Button(self.root, text="Quay lại", command=self.hien_thi_man_hinh_chinh).grid(row=len(danh_sach) + 1, column=0)

    # --- XÓA XE ---
    def hien_thi_man_hinh_xoa_xe(self):
        self.xoa_man_hinh()
        tk.Label(self.root, text="Nhập ID xe:").grid(row=0, column=0)
        self.id_xoa_entry = tk.Entry(self.root)
        self.id_xoa_entry.grid(row=0, column=1)

        tk.Button(self.root, text="Xóa", command=self.xoa_xe).grid(row=1, column=0, columnspan=2)
        tk.Button(self.root, text="Quay lại", command=self.hien_thi_man_hinh_chinh).grid(row=2, column=0, columnspan=2)

    def xoa_xe(self):
        id_xe = self.id_xoa_entry.get().strip()
        if not id_xe:
            messagebox.showerror("Lỗi", "Vui lòng nhập ID xe!")
            return

        ket_qua = Xe.xoa_xe(id_xe)
        messagebox.showinfo("Thông báo", ket_qua)
        self.hien_thi_man_hinh_chinh()

    # --- TÌM KIẾM ---
    def hien_thi_man_hinh_tim_kiem_xe(self):
        self.xoa_man_hinh()
        tk.Label(self.root, text="Nhập ID/biển số:").grid(row=0, column=0)
        self.tim_kiem_entry = tk.Entry(self.root)
        self.tim_kiem_entry.grid(row=0, column=1)

        tk.Button(self.root, text="Tìm kiếm", command=self.tim_kiem).grid(row=1, column=0, columnspan=2)
        tk.Button(self.root, text="Quay lại", command=self.hien_thi_man_hinh_chinh).grid(row=2, column=0, columnspan=2)

    def tim_kiem(self):
        keyword = self.tim_kiem_entry.get().strip()
        if not keyword:
            messagebox.showerror("Lỗi", "Vui lòng nhập từ khóa!")
            return

        xe = Xe.tim_kiem_xe(keyword)
        if xe:
            info = f"""Thông tin xe:
ID: {xe[0]}
Biển số: {xe[1]}
Loại xe: {xe[2]}
Thời gian vào: {xe[3]}
Loại thanh toán: {xe[5]}"""
            messagebox.showinfo("Kết quả", info)
        else:
            messagebox.showinfo("Thông báo", "Không tìm thấy xe!")

    # --- VÉ THÁNG ---
    def hien_thi_man_hinh_quan_ly_ve_thang(self):
        self.xoa_man_hinh()
        tk.Button(self.root, text="Thêm vé", command=self.hien_thi_them_ve_thang).grid(row=0, column=0)
        tk.Button(self.root, text="DS vé tháng", command=self.hien_thi_ds_ve_thang).grid(row=0, column=1)
        tk.Button(self.root, text="Sửa vé", command=self.hien_thi_sua_ve_thang).grid(row=0, column=2)
        tk.Button(self.root, text="Xóa vé", command=self.hien_thi_xoa_ve_thang).grid(row=0, column=3)
        tk.Button(self.root, text="Quay lại", command=self.hien_thi_man_hinh_chinh).grid(row=0, column=4)

    def hien_thi_them_ve_thang(self):
        self.xoa_man_hinh()
        tk.Label(self.root, text="Biển số:").grid(row=0, column=0)
        self.ve_bien_so_entry = tk.Entry(self.root)
        self.ve_bien_so_entry.grid(row=0, column=1)

        tk.Label(self.root, text="Loại xe:").grid(row=1, column=0)
        self.ve_loai_xe_combobox = ttk.Combobox(self.root, values=["Xe máy", "Ô tô"])
        self.ve_loai_xe_combobox.grid(row=1, column=1)
        self.ve_loai_xe_combobox.current(0)

        tk.Label(self.root, text="Thời hạn (tháng):").grid(row=2, column=0)
        self.ve_thoi_han_entry = tk.Entry(self.root)
        self.ve_thoi_han_entry.grid(row=2, column=1)

        tk.Button(self.root, text="Lưu", command=self.luu_ve_thang).grid(row=3, column=0, columnspan=2)
        tk.Button(self.root, text="Quay lại", command=self.hien_thi_man_hinh_quan_ly_ve_thang).grid(row=4, column=0, columnspan=2)

    def luu_ve_thang(self):
        bien_so = self.ve_bien_so_entry.get().strip()
        loai_xe = self.ve_loai_xe_combobox.get()
        thoi_han = self.ve_thoi_han_entry.get().strip()

        if not bien_so or not thoi_han.isdigit():
            messagebox.showerror("Lỗi", "Dữ liệu không hợp lệ!")
            return

        if VeThang.tao_ve_thang(bien_so, loai_xe, int(thoi_han)):
            # Xuất file Excel
            self.xuat_file_excel(bien_so, loai_xe, int(thoi_han))
            messagebox.showinfo("Thành công", "Thêm vé thành công và đã xuất file Excel!")
        else:
            messagebox.showerror("Lỗi", "Biển số đã tồn tại!")
        self.hien_thi_man_hinh_quan_ly_ve_thang()

    def xuat_file_excel(self, bien_so, loai_xe, thoi_han):
        # Tạo file Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Vé tháng"

        # Thêm thông tin vào file Excel
        ws['A1'] = "Bãi xe Trọng Đạo"
        ws['A2'] = "Thời hạn vé (tháng):"
        ws['B2'] = thoi_han
        ws['A3'] = "Mã vé:"
        ws['B3'] = VeThang.lay_danh_sach_ve_thang()[-1][0]  # Lấy ID vé mới nhất
        ws['A4'] = "Biển số xe:"
        ws['B4'] = bien_so
        ws['A5'] = "Loại xe:"
        ws['B5'] = loai_xe
        ws['A6'] = "Tài khoản cấp vé:"
        ws['B6'] = self.ten_dang_nhap_entry.get()  # Lấy tên tài khoản từ ô đăng nhập

        # Lưu file Excel
        ten_file = f"ve_thang_{bien_so}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(ten_file)
        messagebox.showinfo("Thông báo", f"Đã xuất file Excel: {ten_file}")

    def hien_thi_ds_ve_thang(self):
        self.xoa_man_hinh()
        danh_sach = VeThang.lay_danh_sach_ve_thang()

        if not danh_sach:
            tk.Label(self.root, text="Không có vé tháng").pack()
        else:
            for idx, ve in enumerate(danh_sach):
                tk.Label(self.root, text=f"ID: {ve[0]} | Biển: {ve[1]} | Loại: {ve[2]} | HSD: {ve[4]}").grid(row=idx, column=0)

        tk.Button(self.root, text="Quay lại", command=self.hien_thi_man_hinh_quan_ly_ve_thang).grid(row=len(danh_sach) + 1, column=0)

    def hien_thi_sua_ve_thang(self):
        self.xoa_man_hinh()
        tk.Label(self.root, text="ID vé cần sửa:").grid(row=0, column=0)
        self.sua_id_ve_entry = tk.Entry(self.root)
        self.sua_id_ve_entry.grid(row=0, column=1)

        tk.Label(self.root, text="Biển số mới:").grid(row=1, column=0)
        self.sua_bien_so_entry = tk.Entry(self.root)
        self.sua_bien_so_entry.grid(row=1, column=1)

        tk.Label(self.root, text="Loại xe mới:").grid(row=2, column=0)
        self.sua_loai_xe_combobox = ttk.Combobox(self.root, values=["Xe máy", "Ô tô"])
        self.sua_loai_xe_combobox.grid(row=2, column=1)
        self.sua_loai_xe_combobox.current(0)

        tk.Label(self.root, text="Thời hạn mới (tháng):").grid(row=3, column=0)
        self.sua_thoi_han_entry = tk.Entry(self.root)
        self.sua_thoi_han_entry.grid(row=3, column=1)

        tk.Button(self.root, text="Lưu", command=self.luu_sua_ve_thang).grid(row=4, column=0, columnspan=2)
        tk.Button(self.root, text="Quay lại", command=self.hien_thi_man_hinh_quan_ly_ve_thang).grid(row=5, column=0, columnspan=2)

    def luu_sua_ve_thang(self):
        id_ve = self.sua_id_ve_entry.get().strip()
        bien_so_moi = self.sua_bien_so_entry.get().strip()
        loai_xe_moi = self.sua_loai_xe_combobox.get()
        thoi_han_moi = self.sua_thoi_han_entry.get().strip()

        if not id_ve.isdigit() or (bien_so_moi and not thoi_han_moi.isdigit()):
            messagebox.showerror("Lỗi", "Dữ liệu không hợp lệ!")
            return

        thoi_han_moi = int(thoi_han_moi) if thoi_han_moi else None
        if VeThang.sua_ve_thang(int(id_ve), bien_so_moi, loai_xe_moi, thoi_han_moi):
            messagebox.showinfo("Thành công", "Sửa vé thành công!")
        else:
            messagebox.showerror("Lỗi", "Không thể sửa vé!")
        self.hien_thi_man_hinh_quan_ly_ve_thang()

    def hien_thi_xoa_ve_thang(self):
        self.xoa_man_hinh()
        tk.Label(self.root, text="ID vé cần xóa:").grid(row=0, column=0)
        self.xoa_id_ve_entry = tk.Entry(self.root)
        self.xoa_id_ve_entry.grid(row=0, column=1)

        tk.Button(self.root, text="Xóa", command=self.xoa_ve_thang).grid(row=1, column=0, columnspan=2)
        tk.Button(self.root, text="Quay lại", command=self.hien_thi_man_hinh_quan_ly_ve_thang).grid(row=2, column=0, columnspan=2)

    def xoa_ve_thang(self):
        id_ve = self.xoa_id_ve_entry.get().strip()
        if not id_ve.isdigit():
            messagebox.showerror("Lỗi", "ID không hợp lệ!")
            return

        VeThang.xoa_ve_thang(int(id_ve))
        messagebox.showinfo("Thành công", "Xóa vé thành công!")
        self.hien_thi_man_hinh_quan_ly_ve_thang()

    def hien_thi_man_hinh_lich_su(self):
        self.xoa_man_hinh()

        frame_vao = tk.Frame(self.root)
        frame_vao.grid(row=0, column=0, padx=10, pady=10)

        tk.Label(frame_vao, text="Lịch sử vào", font=("Arial", 14, "bold")).grid(row=0, column=0)
        lich_su_vao = Xe.lay_lich_su_vao()
        if not lich_su_vao:
            tk.Label(frame_vao, text="Không có dữ liệu").grid(row=1, column=0)
        else:
            for i, xe in enumerate(lich_su_vao):
                tk.Label(frame_vao, text=f"ID: {xe[0]} | Biển: {xe[1]} | Loại: {xe[2]} | Vào: {xe[3]}").grid(row=i+1, column=0)
        frame_ra = tk.Frame(self.root)
        frame_ra.grid(row=0, column=1, padx=10, pady=10)

        tk.Label(frame_ra, text="Lịch sử ra", font=("Arial", 14, "bold")).grid(row=0, column=0)
        lich_su_ra = Xe.lay_lich_su_ra()
        if not lich_su_ra:
            tk.Label(frame_ra, text="Không có dữ liệu").grid(row=1, column=0)
        else:
            for i, xe in enumerate(lich_su_ra):
                tk.Label(frame_ra, text=f"ID: {xe[0]} | Biển: {xe[1]} | Loại: {xe[2]} | Ra: {xe[3]}").grid(row=i+1, column=0)

        tk.Button(self.root, text="Quay lại", command=self.hien_thi_man_hinh_chinh).grid(row=1, column=0, columnspan=2)